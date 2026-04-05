"""KOLAS 양식 불확도 예산표 엑셀 출력.

KOLAS-G-002 "측정결과의 불확도 추정 및 표현을 위한 지침"에 정확히 맞는
불확도 예산표를 .xlsx 파일로 생성.

표준 형식:
- 헤더 정보: 기관명, 측정량, 단위, 날짜 등
- 예산표 칼럼: 성분명, 기호, 평가유형, 분포, 표준불확도, 감도계수, 기여도, 기여율, 자유도
- 요약: 합성표준불확도, 유효자유도, 포함인자, 확장불확도, 신뢰수준
- 바닥글: 표준불확도 표현 문구
"""

from __future__ import annotations

import math
from datetime import datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from ..core.gum import GUMResult


def export_budget_excel(
    result: GUMResult,
    filepath: Optional[str] = None,
    organization: str = "",
    date_str: Optional[str] = None,
) -> BytesIO:
    """불확도 예산표를 KOLAS 양식 엑셀로 출력.

    Args:
        result: GUMResult 계산 결과
        filepath: 저장 경로 (None이면 BytesIO 반환)
        organization: 기관명 (헤더에 표시할 조직 이름)
        date_str: 측정일자 (ISO 형식, 기본값: 현재 날짜)

    Returns:
        BytesIO 객체 (Streamlit 다운로드용)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "불확도 예산표"

    # 기본 설정
    if date_str is None:
        date_str = datetime.now().strftime("%Y-%m-%d")

    # 스타일 정의
    header_font = Font(name="맑은 고딕", bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    cell_font = Font(name="맑은 고딕", size=10)
    title_font = Font(name="맑은 고딕", bold=True, size=14)
    subtitle_font = Font(name="맑은 고딕", bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")

    # === KOLAS-G-002 헤더 정보 (Row 1-4) ===

    # 제목
    ws.merge_cells("A1:J1")
    cell = ws["A1"]
    cell.value = "측정불확도 예산표 (Uncertainty Budget)"
    cell.font = title_font
    cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 25

    # KOLAS 헤더 정보 (기관명, 측정량, 단위, 날짜)
    row = 2

    # 기관명
    ws[f"A{row}"].value = "기관명:"
    ws[f"A{row}"].font = subtitle_font
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"].value = organization if organization else "(조직명)"
    ws[f"B{row}"].font = cell_font

    # 측정량 (상단)
    ws[f"E{row}"].value = "측정량:"
    ws[f"E{row}"].font = subtitle_font
    ws.merge_cells(f"F{row}:J{row}")
    ws[f"F{row}"].value = result.measurand_name
    ws[f"F{row}"].font = cell_font

    row += 1

    # 단위
    ws[f"A{row}"].value = "단위:"
    ws[f"A{row}"].font = subtitle_font
    ws.merge_cells(f"B{row}:D{row}")
    ws[f"B{row}"].value = result.measurand_unit if result.measurand_unit else "-"
    ws[f"B{row}"].font = cell_font

    # 측정일자
    ws[f"E{row}"].value = "측정일자:"
    ws[f"E{row}"].font = subtitle_font
    ws.merge_cells(f"F{row}:J{row}")
    ws[f"F{row}"].value = date_str
    ws[f"F{row}"].font = cell_font

    row += 1

    # 측정 모델
    ws[f"A{row}"].value = "측정 모델:"
    ws[f"A{row}"].font = subtitle_font
    ws.merge_cells(f"B{row}:J{row}")
    ws[f"B{row}"].value = f"Y = {result.model_expression}"
    ws[f"B{row}"].font = cell_font

    # === KOLAS-G-002 예산표 헤더 (Row 6) ===
    header_row = 6

    # KOLAS-G-002 표준 칼럼 구성
    headers = [
        "불확도 성분",
        "기호",
        "평가유형",
        "확률분포",
        "표준불확도\nu(xi)",
        "감도계수\nci",
        "불확도 기여\n|ci|·u(xi)",
        "기여율\n(%)",
        "자유도\nνi",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col)
        cell.value = header
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # 열 너비 설정 (자유도 칼럼 추가)
    widths = [20, 10, 10, 12, 14, 14, 14, 10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[header_row].height = 30

    # === 데이터 행 (예산표) ===
    data_start_row = header_row + 1

    for row_offset, comp in enumerate(result.components):
        row_idx = data_start_row + row_offset
        src = comp.source

        # 분포명 (B형만)
        dist_name = src.distribution.value if src.eval_type == "B" else "-"

        # 자유도 표현 (무한대는 ∞, 유한값은 정수)
        dof_str = "∞" if math.isinf(comp.dof) else f"{comp.dof:.1f}"

        # 평가유형 표현 (A형/B형만, 자유도는 별도 칼럼)
        eval_type_str = f"{src.eval_type}형"

        values = [
            src.name,                                    # 불확도 성분
            src.symbol,                                 # 기호
            eval_type_str,                              # 평가유형 (A형/B형)
            dist_name,                                  # 확률분포
            comp.std_uncertainty,                       # 표준불확도 (숫자)
            comp.sensitivity_coeff,                     # 감도계수 (숫자)
            comp.contribution,                          # 불확도 기여 (숫자)
            comp.percent_contribution,                  # 기여율 (%)
            dof_str,                                    # 자유도
        ]

        # 셀 속성
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.value = val
            cell.font = cell_font
            cell.border = thin_border

            # 열별 정렬과 숫자 형식
            if col <= 4:  # 처음 4개 열은 텍스트 (좌측 정렬)
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif col in (5, 6, 7):  # 숫자 열: 과학 표기법
                cell.alignment = center
                # 과학 표기법: e.g., 1.23E-04
                if isinstance(val, (int, float)):
                    cell.number_format = '0.00E+00'
            elif col == 8:  # 기여율: 소수점 1자리
                cell.alignment = center
                if isinstance(val, (int, float)):
                    cell.number_format = '0.0'
            else:  # 자유도: 텍스트 중앙
                cell.alignment = center

        # 행 높이 자동 조정
        ws.row_dimensions[row_idx].height = 20

    # === KOLAS-G-002 요약 섹션 (Summary) ===
    summary_row = data_start_row + len(result.components) + 2

    summary_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
    summary_font_label = Font(name="맑은 고딕", bold=True, size=11)
    summary_font_value = Font(name="맑은 고딕", bold=True, size=11, color="CC0000")

    # 1. 합성표준불확도 uc(y)
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    label_cell = ws[f"A{summary_row}"]
    label_cell.value = "합성표준불확도 uc(y)"
    label_cell.font = summary_font_label
    label_cell.fill = summary_fill
    label_cell.border = thin_border
    label_cell.alignment = left_align

    ws.merge_cells(f"E{summary_row}:I{summary_row}")
    value_cell = ws[f"E{summary_row}"]
    value_cell.value = result.combined_uncertainty
    value_cell.font = summary_font_value
    value_cell.fill = summary_fill
    value_cell.border = thin_border
    value_cell.alignment = center
    value_cell.number_format = '0.00E+00'

    summary_row += 1

    # 2. 유효자유도 νeff
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    label_cell = ws[f"A{summary_row}"]
    label_cell.value = "유효자유도 νeff"
    label_cell.font = summary_font_label
    label_cell.fill = summary_fill
    label_cell.border = thin_border
    label_cell.alignment = left_align

    ws.merge_cells(f"E{summary_row}:I{summary_row}")
    value_cell = ws[f"E{summary_row}"]
    dof_display = "∞" if math.isinf(result.effective_dof) else f"{result.effective_dof:.1f}"
    value_cell.value = dof_display
    value_cell.font = cell_font
    value_cell.fill = summary_fill
    value_cell.border = thin_border
    value_cell.alignment = center

    summary_row += 1

    # 3. 포함인자 k
    conf_pct = result.confidence_level * 100
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    label_cell = ws[f"A{summary_row}"]
    label_cell.value = f"포함인자 k (신뢰수준 {conf_pct:.0f}%)"
    label_cell.font = summary_font_label
    label_cell.fill = summary_fill
    label_cell.border = thin_border
    label_cell.alignment = left_align

    ws.merge_cells(f"E{summary_row}:I{summary_row}")
    value_cell = ws[f"E{summary_row}"]
    value_cell.value = result.coverage_factor
    value_cell.font = cell_font
    value_cell.fill = summary_fill
    value_cell.border = thin_border
    value_cell.alignment = center
    value_cell.number_format = '0.00'

    summary_row += 1

    # 4. 확장불확도 U = k · uc(y) (강조)
    ws.merge_cells(f"A{summary_row}:D{summary_row}")
    label_cell = ws[f"A{summary_row}"]
    label_cell.value = "확장불확도 U = k · uc(y)"
    label_cell.font = Font(name="맑은 고딕", bold=True, size=12)
    label_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    label_cell.border = thin_border
    label_cell.alignment = left_align

    ws.merge_cells(f"E{summary_row}:I{summary_row}")
    value_cell = ws[f"E{summary_row}"]
    value_cell.value = result.expanded_uncertainty
    value_cell.font = Font(name="맑은 고딕", bold=True, size=12, color="CC0000")
    value_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    value_cell.border = thin_border
    value_cell.alignment = center
    value_cell.number_format = '0.00E+00'

    ws.row_dimensions[summary_row].height = 22

    # === 바닥글: 불확도 표현 문구 ===
    summary_row += 2
    ws.merge_cells(f"A{summary_row}:I{summary_row}")
    footer_cell = ws[f"A{summary_row}"]
    footer_cell.value = result.uncertainty_statement()
    footer_cell.font = Font(name="맑은 고딕", italic=True, size=10)
    footer_cell.alignment = Alignment(horizontal="center", vertical="center")
    footer_cell.border = thin_border
    ws.row_dimensions[summary_row].height = 20

    # 저장
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    if filepath:
        with open(filepath, "wb") as f:
            f.write(buf.getvalue())
        buf.seek(0)

    return buf
